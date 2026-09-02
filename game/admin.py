"""
Admin configuration for game app
"""
from django.contrib import admin
from django.http import HttpResponse
from django.urls import reverse
from django.utils.html import format_html
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

from .models import GameSpin, LeapWheel, LeapWheelPrize, LeapWheelSpin


class LeapWheelPrizeInline(admin.TabularInline):
    model = LeapWheelPrize
    extra = 0
    max_num = 7
    can_delete = False
    fields = ('name', 'logo', 'logo_preview', 'total_quantity', 'remaining_quantity', 'is_active', 'order')
    readonly_fields = ('name', 'logo_preview', 'order')
    ordering = ('order', 'id')

    def has_add_permission(self, request, obj=None):
        if obj and obj.prizes.count() >= 7:
            return False
        return super().has_add_permission(request, obj)

    @admin.display(description='معاينة الشعار')
    def logo_preview(self, obj):
        if obj.logo:
            return format_html('<img src="{}" style="max-height:40px;max-width:80px;">', obj.logo.url)
        return '—'


@admin.register(LeapWheel)
class LeapWheelAdmin(admin.ModelAdmin):
    list_display = ('title', 'enabled_badge', 'prizes_stock', 'public_link', 'updated_at')
    list_display_links = ('title',)
    fields = (
        'title', 'is_enabled',
        'sponsor_1_name', 'sponsor_1_logo',
        'sponsor_2_name', 'sponsor_2_logo',
        'public_link_readonly', 'updated_at',
    )
    readonly_fields = ('public_link_readonly', 'updated_at')
    inlines = [LeapWheelPrizeInline]

    def has_add_permission(self, request):
        if LeapWheel.objects.exists():
            return False
        return super().has_add_permission(request)

    def has_delete_permission(self, request, obj=None):
        return False

    def changelist_view(self, request, extra_context=None):
        wheel = LeapWheel.get_instance()
        wheel.ensure_default_prizes()
        return super().changelist_view(request, extra_context)

    def change_view(self, request, object_id, form_url='', extra_context=None):
        wheel = LeapWheel.get_instance()
        wheel.ensure_default_prizes()
        if str(object_id) != '1':
            from django.shortcuts import redirect
            return redirect(reverse('admin:game_leapwheel_change', args=[1]))
        return super().change_view(request, object_id, form_url, extra_context)

    @admin.display(description='الحالة')
    def enabled_badge(self, obj):
        if obj.is_enabled:
            return format_html('<span style="color:#10B981;font-weight:700;">● مفعّلة</span>')
        return format_html('<span style="color:#EF4444;font-weight:700;">● متوقفة</span>')

    @admin.display(description='المخزون')
    def prizes_stock(self, obj):
        from django.db.models import Sum
        agg = obj.prizes.filter(is_active=True).aggregate(
            total=Sum('total_quantity'),
            remaining=Sum('remaining_quantity'),
        )
        remaining = agg.get('remaining') or 0
        total = agg.get('total') or 0
        return f"{remaining} / {total}"

    @admin.display(description='رابط العجلة')
    def public_link(self, obj):
        url = reverse('game:leap_wheel')
        return format_html('<a href="{}" target="_blank">{}</a>', url, url)

    @admin.display(description='رابط الصفحة العامة')
    def public_link_readonly(self, obj):
        url = reverse('game:leap_wheel')
        full = self.request.build_absolute_uri(url) if hasattr(self, 'request') else url
        return format_html(
            '<a href="{}" target="_blank" style="font-weight:700;">{}</a>'
            '<p style="color:#6B7280;margin:8px 0 0;">'
            'الجوائز السبعة ثابتة على العجلة — أضف الكمية وارفع شعارات الشركاء من هنا'
            '</p>',
            url, full
        )

    def get_form(self, request, obj=None, **kwargs):
        self.request = request
        return super().get_form(request, obj, **kwargs)


@admin.register(LeapWheelSpin)
class LeapWheelSpinAdmin(admin.ModelAdmin):
    list_display = ('visitor_name', 'visitor_phone', 'prize_name', 'created_at')
    list_filter = ('created_at',)
    search_fields = ('visitor_name', 'visitor_phone', 'prize_name')
    readonly_fields = (
        'leap_wheel', 'prize', 'visitor_name', 'visitor_phone',
        'prize_name', 'ip_address', 'user_agent', 'created_at'
    )
    actions = ['export_to_excel']

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def changelist_view(self, request, extra_context=None):
        if 'action' in request.POST and request.POST['action'] == 'export_to_excel':
            selected_ids = request.POST.getlist('_selected_action')
            queryset = self.get_queryset(request)
            if selected_ids:
                queryset = queryset.filter(pk__in=selected_ids)
            return self.export_to_excel(request, queryset)

        extra_context = extra_context or {}
        extra_context['show_export_button'] = True
        extra_context['export_action_name'] = 'export_to_excel'
        return super().changelist_view(request, extra_context)

    def export_to_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "دورات LEAP"
        headers = ['ID', 'الاسم', 'الجوال', 'الجائزة', 'التاريخ']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6D28D9", end_color="6D28D9", fill_type="solid")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_num, spin in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=spin.id)
            ws.cell(row=row_num, column=2, value=spin.visitor_name)
            ws.cell(row=row_num, column=3, value=spin.visitor_phone)
            ws.cell(row=row_num, column=4, value=spin.prize_name)
            created = spin.created_at.strftime('%Y-%m-%d %H:%M') if spin.created_at else '-'
            ws.cell(row=row_num, column=5, value=created)

        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 22

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'leap_spins_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    export_to_excel.short_description = "📊 تصدير دورات LEAP إلى Excel"


@admin.register(GameSpin)
class GameSpinAdmin(admin.ModelAdmin):
    list_display = [
        'visitor_name',
        'visitor_phone',
        'company',
        'prize',
        'won',
        'created_at',
        'ip_address'
    ]
    list_filter = ['won', 'company', 'created_at']
    search_fields = ['visitor_name', 'visitor_phone', 'prize', 'company__name']
    readonly_fields = ['created_at', 'session_id', 'ip_address', 'user_agent']
    actions = ['export_to_excel']

    fieldsets = (
        ('معلومات الدورة', {
            'fields': ('company', 'visitor_name', 'visitor_phone', 'prize', 'won')
        }),
        ('معلومات تقنية', {
            'fields': ('session_id', 'ip_address', 'user_agent', 'created_at'),
            'classes': ('collapse',)
        }),
    )

    def get_queryset(self, request):
        return super().get_queryset(request).select_related('company')

    def changelist_view(self, request, extra_context=None):
        if 'action' in request.POST and request.POST['action'] == 'export_to_excel':
            selected_ids = request.POST.getlist('_selected_action')
            if selected_ids:
                queryset = self.get_queryset(request).filter(pk__in=selected_ids)
            else:
                queryset = self.get_queryset(request)
            return self.export_to_excel(request, queryset)

        extra_context = extra_context or {}
        extra_context['show_export_button'] = True
        extra_context['export_action_name'] = 'export_to_excel'
        return super().changelist_view(request, extra_context)

    def export_to_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "دورات الألعاب"

        headers = [
            'ID', 'الشركة', 'اسم الزائر', 'رقم الجوال', 'الجائزة',
            'فاز', 'تاريخ الدورة', 'عنوان IP', 'المتصفح',
        ]

        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="6A3FA0", end_color="6A3FA0", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        row_num = 2
        for spin in queryset:
            ws.cell(row=row_num, column=1, value=spin.id)
            ws.cell(row=row_num, column=2, value=getattr(spin.company, 'name', '-'))
            ws.cell(row=row_num, column=3, value=spin.visitor_name)
            ws.cell(row=row_num, column=4, value=spin.visitor_phone or '-')
            ws.cell(row=row_num, column=5, value=spin.prize)
            ws.cell(row=row_num, column=6, value='نعم' if spin.won else 'لا')
            created_val = spin.created_at
            if created_val and hasattr(created_val, 'strftime'):
                ws.cell(row=row_num, column=7, value=created_val.strftime('%Y-%m-%d %H:%M'))
            else:
                ws.cell(row=row_num, column=7, value=str(created_val) if created_val else '-')
            ws.cell(row=row_num, column=8, value=spin.ip_address or '-')
            ws.cell(row=row_num, column=9, value=spin.user_agent or '-')
            row_num += 1

        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'دورات_الألعاب_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
        wb.save(response)
        return response

    export_to_excel.short_description = "📊 تصدير دورات الألعاب إلى Excel"
